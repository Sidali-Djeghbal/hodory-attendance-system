from typing import List, Optional, Dict, Any
from sqlmodel import Session, select
from fastapi import HTTPException, status
from datetime import datetime
import os

from ..models.report import Report
from ..models.admin import Admin

# Directory for saving reports
REPORTS_DIR = "uploads/reports"


class ReportController:
    """
    Report Controller - Generates and saves reports to uploads/reports/
    
    Workflow:
        1. AdminController calls generate_report() with attendance data
        2. ReportController creates PDF and Excel files
        3. Saves file URLs in Report model
        4. Returns report info
    """
    
    def __init__(self, session: Session):
        self.session = session
        os.makedirs(REPORTS_DIR, exist_ok=True)
    
    def generate_report(
        self,
        admin_id: int,
        report_data: Dict[str, Any],
        period_start: datetime,
        period_end: datetime
    ) -> Dict[str, Any]:
        """
        Generate report with PDF and Excel files saved to uploads/reports/
        
        Args:
            admin_id: ID of the admin
            report_data: Data from AdminController.generate_report()
            period_start: Report period start
            period_end: Report period end
            
        Returns:
            dict: Report info with file URLs
        """
        # Verify admin exists
        admin = self.session.get(Admin, admin_id)
        if not admin:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Admin with ID {admin_id} not found"
            )
        
        # Build content summary
        period_summary = report_data.get("period_summary", {})
        content = (
            f"Attendance Report | "
            f"Total: {period_summary.get('total_attendance_records', 0)} | "
            f"Present: {period_summary.get('present', 0)} | "
            f"Absent: {period_summary.get('absent', 0)} | "
            f"Rate: {period_summary.get('attendance_rate', 0)}%"
        )
        
        # Create report record
        report = Report(
            admin_id=admin_id,
            content=content,
            period_start=period_start,
            period_end=period_end,
            generated_date=datetime.utcnow()
        )
        self.session.add(report)
        self.session.commit()
        self.session.refresh(report)
        
        # Generate file names
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        pdf_filename = f"report_{report.id}_{timestamp}.pdf"
        excel_filename = f"report_{report.id}_{timestamp}.xlsx"
        pdf_path = os.path.join(REPORTS_DIR, pdf_filename)
        excel_path = os.path.join(REPORTS_DIR, excel_filename)
        
        # Generate PDF
        try:
            self._create_pdf(report, report_data, pdf_path)
            report.pdf_url = pdf_path
        except Exception as e:
            print(f"PDF generation error: {e}")
            report.pdf_url = None
        
        # Generate Excel
        try:
            self._create_excel(report, report_data, excel_path)
            report.excel_url = excel_path
        except Exception as e:
            print(f"Excel generation error: {e}")
            report.excel_url = None
        
        # Update report with file URLs
        self.session.add(report)
        self.session.commit()
        self.session.refresh(report)
        
        return {
            "report_id": report.id,
            "admin_id": admin_id,
            "generated_date": report.generated_date,
            "period_start": period_start,
            "period_end": period_end,
            "content": content,
            "pdf_url": report.pdf_url,
            "excel_url": report.excel_url,
            "summary": period_summary
        }
    
    def _create_pdf(self, report: Report, data: Dict[str, Any], file_path: str):
        """Create PDF report file"""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=20)
        heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, spaceAfter=10)
        
        # Title
        elements.append(Paragraph(f"Attendance Report #{report.id}", title_style))
        elements.append(Paragraph(f"Generated: {report.generated_date}", styles['Normal']))
        elements.append(Paragraph(f"Period: {report.period_start} to {report.period_end}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Summary Table
        elements.append(Paragraph("Summary", heading_style))
        summary = data.get("period_summary", {})
        summary_data = [
            ['Metric', 'Value'],
            ['Total Records', str(summary.get('total_attendance_records', 0))],
            ['Present', str(summary.get('present', 0))],
            ['Absent', str(summary.get('absent', 0))],
            ['Justified', str(summary.get('justified', 0))],
            ['Excluded', str(summary.get('excluded', 0))],
            ['Attendance Rate', f"{summary.get('attendance_rate', 0)}%"]
        ]
        
        table = Table(summary_data, colWidths=[200, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D6DCE4')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 30))
        
        # Levels
        for level in data.get("levels", []):
            elements.append(Paragraph(f"Level: {level.get('name', 'Unknown')}", heading_style))
            
            # Students
            students = level.get("students", [])
            if students:
                student_data = [['Student', 'Email', 'Enrollments']]
                for s in students[:20]:
                    student_data.append([
                        s.get('name', 'Unknown'),
                        s.get('email', 'N/A'),
                        str(len(s.get('enrollments', [])))
                    ])
                
                st_table = Table(student_data, colWidths=[150, 200, 80])
                st_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
                ]))
                elements.append(st_table)
            
            elements.append(Spacer(1, 20))
        
        doc.build(elements)
    
    def _create_excel(self, report: Report, data: Dict[str, Any], file_path: str):
        """Create Excel report file"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        ws['A1'] = f"Attendance Report #{report.id}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A3'] = "Generated:"
        ws['B3'] = str(report.generated_date)
        ws['A4'] = "Period:"
        ws['B4'] = f"{report.period_start} to {report.period_end}"
        
        summary = data.get("period_summary", {})
        ws['A6'] = "Metric"
        ws['B6'] = "Value"
        ws['A6'].fill = header_fill
        ws['A6'].font = header_font
        ws['B6'].fill = header_fill
        ws['B6'].font = header_font
        
        stats = [
            ('Total Records', summary.get('total_attendance_records', 0)),
            ('Present', summary.get('present', 0)),
            ('Absent', summary.get('absent', 0)),
            ('Justified', summary.get('justified', 0)),
            ('Excluded', summary.get('excluded', 0)),
            ('Attendance Rate', f"{summary.get('attendance_rate', 0)}%")
        ]
        for i, (metric, value) in enumerate(stats, 7):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = str(value)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        
        # Students Sheet
        ws_students = wb.create_sheet("Students")
        headers = ['Level', 'Student', 'Email', 'Enrollments']
        for col, h in enumerate(headers, 1):
            cell = ws_students.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for level in data.get("levels", []):
            for student in level.get("students", []):
                ws_students.cell(row=row, column=1, value=level.get('name', 'Unknown'))
                ws_students.cell(row=row, column=2, value=student.get('name', 'Unknown'))
                ws_students.cell(row=row, column=3, value=student.get('email', 'N/A'))
                ws_students.cell(row=row, column=4, value=len(student.get('enrollments', [])))
                row += 1
        
        for col in ['A', 'B', 'C', 'D']:
            ws_students.column_dimensions[col].width = 25
        
        # Attendance Sheet
        ws_att = wb.create_sheet("Attendance")
        headers = ['Level', 'Student', 'Module', 'Date', 'Status']
        for col, h in enumerate(headers, 1):
            cell = ws_att.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for level in data.get("levels", []):
            for student in level.get("students", []):
                for enrollment in student.get("enrollments", []):
                    for record in enrollment.get("attendance_records", []):
                        ws_att.cell(row=row, column=1, value=level.get('name', 'Unknown'))
                        ws_att.cell(row=row, column=2, value=student.get('name', 'Unknown'))
                        ws_att.cell(row=row, column=3, value=enrollment.get('module_name', 'Unknown'))
                        ws_att.cell(row=row, column=4, value=str(record.get('session_date', 'N/A')))
                        ws_att.cell(row=row, column=5, value=record.get('status', 'ABSENT'))
                        row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_att.column_dimensions[col].width = 25
        
        wb.save(file_path)
    
    def get_report_by_id(self, report_id: int) -> Report:
        """Get report by ID"""
        report = self.session.get(Report, report_id)
        if not report:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Report with ID {report_id} not found"
            )
        return report
    
    def get_reports_by_admin(self, admin_id: int, limit: int = 50) -> List[Report]:
        """Get all reports by admin"""
        return self.session.exec(
            select(Report)
            .where(Report.admin_id == admin_id)
            .order_by(Report.generated_date.desc())
            .limit(limit)
        ).all()
    
    def delete_report(self, report_id: int) -> Dict[str, Any]:
        """Delete report and its files"""
        report = self.get_report_by_id(report_id)
        
        # Delete files
        if report.pdf_url and os.path.exists(report.pdf_url):
            os.remove(report.pdf_url)
        if report.excel_url and os.path.exists(report.excel_url):
            os.remove(report.excel_url)
        
        self.session.delete(report)
        self.session.commit()
        
        return {"message": f"Report {report_id} deleted"}
